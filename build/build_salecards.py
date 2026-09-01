# -*- coding: utf-8 -*-
"""26FW 스타일맵.xlsx -> 판매자료 카드 / 시즌 집계표 HTML 생성

  결과물:  share/26fw-sales-cards.html
           share/26fw-season-summary.html
           share/img/{품번}.jpg            (Supabase 에 없는 도식화 이미지)
"""
import base64, json, datetime, os, re, sys
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))                 # .../build
ROOT = os.path.dirname(HERE) if os.path.basename(HERE) == 'build' else HERE
DATA_DIR = os.path.join(ROOT, 'data') if os.path.isdir(os.path.join(ROOT, 'data')) else ROOT
SHARE_DIR = os.path.join(ROOT, 'share')
IMG_DIR = os.path.join(SHARE_DIR, 'img')
os.makedirs(IMG_DIR, exist_ok=True)

# data/ 에서 날짜가 가장 큰 스타일맵 파일 사용 (인자로 경로를 주면 그 파일)
if len(sys.argv) > 1:
    SRC = sys.argv[1]
else:
    cands = sorted(f for f in os.listdir(DATA_DIR)
                   if re.match(r'^\d{6}_26FW 스타일맵.*\.xlsx$', f) and not f.startswith('~$'))
    if not cands:
        raise SystemExit('%s 에 "YYMMDD_26FW 스타일맵.xlsx" 파일이 없습니다.' % DATA_DIR)
    SRC = os.path.join(DATA_DIR, cands[-1])
print('원본:', os.path.basename(SRC))
OUT_CARD = os.path.join(SHARE_DIR, '26fw-sales-cards.html')
OUT_SUM = os.path.join(SHARE_DIR, '26fw-season-summary.html')

# 시트명은 파일마다 '생판재주간-' / '생판재-' 로 다를 수 있어 앞부분만 매칭
def sheet(*keys):
    for name in wb.sheetnames:
        if all(k in name for k in keys):
            return name
    raise KeyError(str(keys))


def C(letter):
    """엑셀 열문자 -> 0-based 인덱스"""
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


# ST 시트 열
ST = dict(season=C('D'), gender=C('E'), item=C('F'), item2=C('H'), theme=C('K'), style=C('M'), name=C('N'),
          vendor=C('P'), real=C('Y'), tag=C('Z'), cost=C('AA'),
          tag_amt=C('BH'), sale_amt=C('BI'), img_code=C('BR'))
# CO 시트 열
CO = dict(style=C('K'), color=C('M'), cname=C('N'), plan=C('R'),
          in_date=C('U'), qty_in=C('V'), out_date=C('Y'),
          w1=C('CE'), w2=C('CD'), total=C('CG'), rate=C('CH'), grp=C('CU'))


def num(v):
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '').replace('%', '').strip())
    except ValueError:
        return None


def txt(v):
    return '' if v is None else str(v).strip()


def date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return '%d/%d' % (v.month, v.day)
    s = txt(v)[:10]
    if len(s) == 10 and s[4] == '-':
        return '%d/%d' % (int(s[5:7]), int(s[8:10]))
    return ''


def dkey(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    return txt(v)[:10]


wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

ST_MAIN, ST_RUN = sheet('생판재', '26FW'), sheet('생판재', '러닝')
CO_MAIN, CO_RUN = sheet('주간판매', '26FW'), sheet('주간판매', '러닝')

# '26FW 러닝스타일 리스트' 시트: 기존품번 <-> 신규러닝품번 상호 매핑 (이미지 폴더 후보용)
pair = {}
if '26FW 러닝스타일 리스트' in wb.sheetnames:
    for r in wb['26FW 러닝스타일 리스트'].iter_rows(min_row=2, values_only=True):
        old, new = txt(r[0]), txt(r[1])
        if old and new:
            pair.setdefault(old, []).append(new)
            pair.setdefault(new, []).append(old)

styles = {}
order = []

for st_sheet, is_main in ((ST_MAIN, True), (ST_RUN, False)):
    ws = wb[st_sheet]
    for r in ws.iter_rows(min_row=4, values_only=True):
        code = txt(r[ST['style']] if len(r) > ST['style'] else '')
        if not code or code == '스타일':
            continue
        season = txt(r[ST['season']]).replace('(러닝)', '').strip()
        tag, cost = num(r[ST['tag']]), num(r[ST['cost']])
        tag_amt, sale_amt = num(r[ST['tag_amt']]), num(r[ST['sale_amt']])
        # 이미지 폴더 후보: 품번 -> 러닝 시트 BR열 품번 -> 러닝 매핑표의 짝 품번 순으로 시도
        img_code = txt(r[ST['img_code']]) if len(r) > ST['img_code'] else ''
        cands = []
        for c in [code, img_code] + pair.get(code, []) + pair.get(img_code, []):
            if c and c not in cands:
                cands.append(c)
        styles.setdefault(code, {
            'style': code,
            'imgCodes': cands,
            'name': txt(r[ST['name']]),
            'season': season,
            'gender': txt(r[ST['gender']]),
            'item': txt(r[ST['item']]),
            'cat2': txt(r[ST['item2']]),   # 아이템(H열) — 집계표 카테고리명과 대응
            # 테마 = 생판재 시트 K열(컨셉). '노르딕,패턴' 도 통째로 하나의 테마다
            'themes': [txt(r[ST['theme']])] if txt(r[ST['theme']]) else [],
            'vendor': txt(r[ST['vendor']]),
            'tag': tag,
            'real': num(r[ST['real']]),
            'cost': cost,
            'mult': round(tag / cost / 1.1, 2) if tag and cost else None,
            'tagAmt': tag_amt,
            'saleAmt': sale_amt,
            'discount': (1 - sale_amt / tag_amt) if tag_amt else None,
            'line': '메인' if is_main else '러닝',
            'colors': [], 'inDate': '', 'outDate': '',
        })
        if code not in order:
            order.append(code)

# 메인/러닝: 26FW 시트(ST/CO)에 존재하면 메인
main_codes = set()
for st_sheet in (ST_MAIN,):
    for r in wb[st_sheet].iter_rows(min_row=4, values_only=True):
        c = txt(r[ST['style']])
        if c and c != '스타일':
            main_codes.add(c)
for r in wb[CO_MAIN].iter_rows(min_row=4, values_only=True):
    c = txt(r[CO['style']])
    if c and c != '스타일':
        main_codes.add(c)

# 파일 연도 (8/1 이후 판매 합산 기준)
_m = re.match(r'^(\d{2})(\d{2})(\d{2})_', os.path.basename(SRC))
FILE_YEAR = 2000 + int(_m.group(1)) if _m else 2026

for co_sheet in (CO_MAIN, CO_RUN):
    ws = wb[co_sheet]
    # 주간 열은 매주 한 칸씩 밀리는 롤링 구조라 라벨(W1=최근 완료 주)로 찾는다
    hdr = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    wk = {}
    w1_start = None
    for i, v in enumerate(hdr):
        m = re.match(r'^W(\d+)\((\d{2})/(\d{2})', txt(v))
        if m:
            wk[int(m.group(1))] = i
            if int(m.group(1)) == 1:
                w1_start = datetime.date(FILE_YEAR, int(m.group(2)), int(m.group(3)))
    w1i, w2i = wk.get(1, C('CE')), wk.get(2, C('CD'))
    # 8/1 이후 판매 = 시작일이 8/1 이상인 주들의 합 (러닝 현재고 판매율용)
    aug_cols = []
    if w1_start:
        # 8/1 이 포함된 주(W5, 07/27~08/02)까지 포함 — 주 끝날짜가 8/1 이상인 주
        aug1 = datetime.date(FILE_YEAR, 8, 1)
        max_n = (w1_start - (aug1 - datetime.timedelta(days=6))).days // 7 + 1
        aug_cols = [idx for n, idx in wk.items() if 1 <= n <= max_n]
    print('%s: 전주 %s / 2주전 %s / 8월이후 주 %d개' % (co_sheet, txt(hdr[w1i]), txt(hdr[w2i]), len(aug_cols)))
    for r in ws.iter_rows(min_row=4, values_only=True):
        code = txt(r[CO['style']] if len(r) > CO['style'] else '')
        if not code or code == '스타일':
            continue
        s = styles.get(code)
        if s is None:
            continue
        s['colors'].append({
            'color': txt(r[CO['color']]),
            'cname': txt(r[CO['cname']]),
            'plan': num(r[CO['plan']]),
            'qtyIn': num(r[CO['qty_in']]),
            'w1': num(r[w1i]) if len(r) > w1i else None,
            'w2': num(r[w2i]) if len(r) > w2i else None,
            'total': num(r[CO['total']]),
            'rate': num(r[CO['rate']]),
            'aug': sum((num(r[i]) or 0) for i in aug_cols if len(r) > i),
        })
        s['augSales'] = (s.get('augSales') or 0) + sum(
            (num(r[i]) or 0) for i in aug_cols if len(r) > i)
        # 입고일/출고일: 칼라별 날짜 중 가장 빠른 날짜
        for key, col in (('_in', 'in_date'), ('_out', 'out_date')):
            k = dkey(r[CO[col]])
            if k and (s.get(key) is None or k < s[key]):
                s[key] = k
                s['inDate' if key == '_in' else 'outDate'] = date(r[CO[col]])
        # 러닝 시트 CU열: 품번 통합용 그룹 품번
        if co_sheet == CO_RUN:
            g = txt(r[CO['grp']]) if len(r) > CO['grp'] else ''
            if g:
                s['grp'] = g


def calc_ttl(colors):
    qin = sum(c['qtyIn'] or 0 for c in colors)
    tot = sum(c['total'] or 0 for c in colors)
    return {'plan': sum(c['plan'] or 0 for c in colors), 'qtyIn': qin,
            'w1': sum(c['w1'] or 0 for c in colors),
            'w2': sum(c['w2'] or 0 for c in colors),
            'total': tot, 'rate': (tot / qin * 100) if qin else None}


for code, s in styles.items():
    s['line'] = '메인' if code in main_codes else '러닝'
    s['arrived'] = '입고' if sum(c['qtyIn'] or 0 for c in s['colors']) > 0 else '미입고'
    s['ttl'] = calc_ttl(s['colors'])

# ---- CU열 기준 품번 통합 (기존품번 + 러닝 신규품번) ----
groups = {}
for code, s in styles.items():
    if s.get('grp'):
        groups.setdefault(s['grp'], []).append(code)

subs = set()
for g, members in groups.items():
    prim = next((m for m in members if m != g), None)
    if not prim or g not in members:
        continue
    p, sub = styles[prim], styles[g]
    # 칼라별 값 합산
    merged = {}
    for c in p['colors'] + sub['colors']:
        m = merged.setdefault(c['color'], {'color': c['color'], 'cname': c['cname'],
                                           'plan': 0, 'qtyIn': 0, 'w1': 0, 'w2': 0, 'total': 0, 'aug': 0})
        for f in ('plan', 'qtyIn', 'w1', 'w2', 'total', 'aug'):
            m[f] += c.get(f) or 0
    for m in merged.values():
        m['rate'] = (m['total'] / m['qtyIn'] * 100) if m['qtyIn'] else None
    p['colors'] = list(merged.values())
    p['ttl'] = calc_ttl(p['colors'])
    p['altStyle'] = g
    for f in ('tagAmt', 'saleAmt'):
        p[f] = (p[f] or 0) + (sub[f] or 0) if (p[f] or sub[f]) else None
    p['discount'] = (1 - p['saleAmt'] / p['tagAmt']) if p['tagAmt'] else None
    p['arrived'] = '입고' if p['ttl']['qtyIn'] > 0 else '미입고'
    for key, fld in (('_in', 'inDate'), ('_out', 'outDate')):
        if sub.get(key) and (p.get(key) is None or sub[key] < p[key]):
            p[key], p[fld] = sub[key], sub[fld]
    for c in sub['imgCodes']:
        if c not in p['imgCodes']:
            p['imgCodes'].append(c)
    p['sub'] = sub                       # 클릭 시 보여줄 CU 품번 단독 카드
    subs.add(g)

for s in styles.values():
    for k in ('_in', '_out', 'grp'):
        s.pop(k, None)

data = [styles[c] for c in order if c not in subs]

# ---- 러닝 현재고: 시점재고 시트(기존품번 M열 -> BO열 '전체') + 신규품번 입고량 ----
point_stock, point_stock_color = {}, {}
stock_sheet = next((n for n in wb.sheetnames if '시점재고' in n), None)
if stock_sheet:
    for r in wb[stock_sheet].iter_rows(min_row=4, values_only=True):
        code = txt(r[C('M')] if len(r) > C('M') else '')
        if not code or code == '스타일':
            continue
        v = num(r[C('BO')]) if len(r) > C('BO') else None
        if v is not None:
            point_stock[code] = point_stock.get(code, 0) + v
            col = txt(r[C('O')] if len(r) > C('O') else '')
            if col:
                k = (code, col)
                point_stock_color[k] = point_stock_color.get(k, 0) + v
    print('시점재고:', len(point_stock), '스타일 /', len(point_stock_color), '컬러')

for s in data:
    if s['line'] != '러닝':
        continue
    sub = s['sub'] if s.get('sub') else s          # 신규품번 (짝 없는 GF7 단독 카드는 자기 자신)
    tt = sub['ttl']
    new_in = tt['qtyIn'] or 0
    codes = [s['style']] + s['imgCodes']
    base_code = next((c for c in codes if c in point_stock), None)
    base = point_stock.get(base_code, 0)
    stock = base + new_in
    # 현재고 기준 판매: 구품번의 8/1 이후 판매 + 신규품번 누계판매
    run_sales = ((s.get('augSales') or 0) if s.get('sub') else 0) + (tt['total'] or 0)
    s['stock'] = stock
    stat = {'base': base, 'addPlan': tt['plan'] or 0, 'newIn': new_in,
            'stock': stock, 'sales': run_sales,
            'rate': (run_sales / stock * 100) if stock else None}
    # 팝업 컬러 표 (신규 기획 컬러만): 기획 = 컬러별 기초재고 + 컬러별 신규입고,
    # 전주/2주전·누계(8/1 이후)는 구+신 합산, 판매율 = 누계 ÷ (기초+입고)
    merged_colors = s['colors']                    # CU 통합본 (w1/w2/aug 포함)
    prows = []
    for sc in sub['colors']:
        mc = next((c for c in merged_colors if c['color'] == sc['color']), None) or sc
        base_c = point_stock_color.get((base_code, sc['color']), 0) if base_code else 0
        in_c = sc['qtyIn'] or 0
        plan_c = base_c + in_c
        tot = mc.get('aug') or 0
        prows.append({'color': sc['color'], 'cname': sc['cname'],
                      'plan': plan_c, 'qtyIn': in_c,
                      'w1': mc.get('w1'), 'w2': mc.get('w2'),
                      'total': tot, 'rate': (tot / plan_c * 100) if plan_c else None})
    s['popup'] = {'stat': stat, 'colors': prows}
    # 본 카드는 신규품번 내용으로 표시 (메인 카드와 동일 레이아웃)
    if s.get('sub'):
        s['ttl'] = sub['ttl']
        s['colors'] = sub['colors']
        for f in ('tagAmt', 'saleAmt', 'discount', 'inDate', 'outDate'):
            s[f] = sub.get(f)
    s['arrived'] = '입고' if new_in > 0 else '미입고'
print('통합 카드:', len(subs), '쌍')
print('styles:', len(data), '/ colors:', sum(len(s['colors']) for s in data))

# ---- 이미지 ----
# 1순위: Supabase 갤러리 URL (파일 용량을 줄이려고 링크로 사용)
# 2순위: 구글 시트에서 뽑아둔 도식화 이미지(base64) 를 카드에 정적으로 삽입
SUPA = 'https://ijawemotwqhlilagqwfs.supabase.co/storage/v1/object/public/product-gallery/'
CACHE = os.path.join(HERE, 'supabase_cache.json')
cache = json.load(open(CACHE, encoding='utf-8')) if os.path.exists(CACHE) else {}


def supa_has(code):
    """Supabase 갤러리에 이미지가 있는지 (결과는 캐시)"""
    if code in cache:
        return cache[code]
    import urllib.request
    try:
        req = urllib.request.Request(SUPA + code + '/cut_0.jpg', method='HEAD')
        cache[code] = urllib.request.urlopen(req, timeout=15).status == 200
    except Exception:
        cache[code] = False
    return cache[code]


import concurrent.futures as _cf
todo = [c for s in styles.values() for c in s['imgCodes'] if c not in cache]
if todo:
    with _cf.ThreadPoolExecutor(16) as ex:
        list(ex.map(supa_has, dict.fromkeys(todo)))
    json.dump(cache, open(CACHE, 'w', encoding='utf-8'))

gs_path = os.path.join(HERE, 'gs_images.json')
gs = json.load(open(gs_path, encoding='utf-8')) if os.path.exists(gs_path) else {}
linked = saved = 0
written = set()
for s in styles.values():
    hit = [c for c in s['imgCodes'] if cache.get(c)]
    if hit:                                   # Supabase 에 있으면 링크로 사용
        s['imgCodes'] = hit + [c for c in s['imgCodes'] if c not in hit]
        linked += 1
        continue
    for c in [s['style']] + s['imgCodes']:    # 없으면 share/img/ 에 파일로 저장
        uri = gs.get(c.upper())
        if not uri:
            continue
        name = c.upper() + '.jpg'
        path = os.path.join(IMG_DIR, name)
        raw = base64.b64decode(uri.split(',', 1)[1])
        if not os.path.exists(path) or open(path, 'rb').read() != raw:
            open(path, 'wb').write(raw)       # 내용이 같으면 다시 쓰지 않음(깃 diff 방지)
        s['img'] = 'img/' + name
        written.add(name)
        saved += 1
        break

# 더 이상 쓰이지 않는 이미지 파일 정리
for f in os.listdir(IMG_DIR):
    if f.lower().endswith('.jpg') and f not in written:
        os.remove(os.path.join(IMG_DIR, f))
        print('  삭제(미사용):', f)

print('images: supabase 링크 %d / 파일 %d장 (share/img) / 카드 %d' % (linked, len(written), len(data)))

# ---------------- '종합' 시트 -> 집계표 데이터 ----------------
# 좌측 블록(신상품) 지표 열. 러닝 블록은 여기서 27칸 오른쪽으로 밀려 있다.
MET = dict(styles='G', sku='H', plan='I', inM='J', inSku='K', qtyIn='L', inRate='M',
           planCost='N', cost='O', tag='Q', mu='S', sales='T', sellRate='U',
           salesTag='V', salesAmt='W', salesCost='X', disc='Y', wkQty='Z', wkAmt='AA')
PREV = dict(qtyIn='AD', sales='AE', sellRate='AF', salesTag='AG', salesAmt='AH', disc='AI')
SUM = {'block1': [], 'newItem': [], 'newCat': [], 'runItem': [], 'runCat': [], 'seasonCat': []}

if '종합' in wb.sheetnames:
    rows = list(wb['종합'].iter_rows(min_row=1, max_row=150, values_only=True))

    def cell(r, col, off=0):
        row = rows[r - 1]
        i = C(col) + off
        return row[i] if len(row) > i else None

    def vals(r, cols, off=0):
        return {k: num(cell(r, c, off)) for k, c in cols.items()}

    # 1) 26FW 시즌별/복종별 + 전년비교 (7~23행)
    season, mode = '', 'new'
    for r in range(7, 24):
        c, d, e = txt(cell(r, 'C')), txt(cell(r, 'D')), txt(cell(r, 'E'))
        if d:
            season = d
        if c == '러닝':
            mode = 'run'
        rec = {'season': '' if mode == 'run' else season, 'item': e, 'mode': mode,
               'cur': vals(r, MET), 'prev': vals(r, PREV)}
        if c == 'TTL':
            rec.update(item='TTL', season='', kind='신상품 TTL' if mode == 'new' else '전체 TTL')
            mode = 'run' if mode == 'new' else mode
        elif e == 'TTL':
            rec['kind'] = ('%s TTL' % season) if mode == 'new' else '러닝 TTL'
        elif e:
            rec['kind'] = '항목'
        else:
            continue
        SUM['block1'].append(rec)

    # 2~4) 신상품/러닝 복종·카테고리별 (30~64행), 러닝은 오른쪽으로 27칸
    for key_item, key_cat, off, cols in (('newItem', 'newCat', 0, ('D', 'E', 'F')),
                                         ('runItem', 'runCat', 27, ('D', 'E', 'F'))):
        gubun = ''
        for r in range(30, 65):
            d, e, f = (txt(cell(r, cols[0], off)), txt(cell(r, cols[1], off)),
                       txt(cell(r, cols[2], off)))
            v = vals(r, MET, off)
            if d.endswith('TOTAL') and 'ACC제외' not in d:
                SUM[key_item].append({'item': d.replace(' TOTAL', '').strip(), **v})
                continue
            if d and not d.endswith('TOTAL'):
                gubun = d
            if f:                                   # 카테고리명이 있는 행만 카테고리 집계
                if any((v.get(k) or 0) for k in ('plan', 'qtyIn', 'sales')):
                    SUM[key_cat].append({'item': gubun, 'cat': f, 'code': e, **v})

    # 5) 시즌(가을/겨울)별 카테고리 상세 — 66행 아래, A열에 시즌이 적혀 있는 블록
    cur_season, gubun = '', ''
    for r in range(66, len(rows) + 1):      # 시트가 150행보다 짧을 수 있다
        a = txt(cell(r, 'A'))
        if a not in ('가을', '겨울'):
            continue
        if a != cur_season:
            cur_season, gubun = a, ''
        d = txt(cell(r, 'D'))
        e, f = txt(cell(r, 'E')), txt(cell(r, 'F'))
        if d and not d.endswith('TOTAL') and 'ACC제외' not in d:
            gubun = d
        if f:
            v = vals(r, MET)
            if any((v.get(k) or 0) for k in ('plan', 'qtyIn', 'sales')):
                SUM['seasonCat'].append({'season': a, 'item': gubun, 'cat': f, 'code': e, **v})

# ---- 직전 스타일맵 대비 증감 ----
# 각 행에 고유 키를 달아 스냅샷과 대조한다. 날짜가 같은 파일을 다시 빌드하면
# 스냅샷을 덮지 않으므로 증감 표기가 그대로 유지된다.
for d in SUM['block1']:
    # kind 는 항목 행이면 모두 '항목' 이라 복종명(item)까지 넣어야 행이 구분된다
    d['_k'] = 'b1|%s|%s|%s|%s' % (d['mode'], d['season'], d['item'], d.get('kind') or '')
for name in ('newItem', 'runItem'):
    for d in SUM[name]:
        d['_k'] = '%s|%s' % (name, d['item'])
for name in ('newCat', 'runCat'):
    for d in SUM[name]:
        d['_k'] = '%s|%s|%s' % (name, d['item'], d['cat'])
for d in SUM['seasonCat']:
    d['_k'] = 'seasonCat|%s|%s|%s' % (d['season'], d['item'], d['cat'])

MET_KEYS = list(MET.keys()) + ['stock']


def row_vals(d):
    src = d['cur'] if 'cur' in d else d
    return {k: src.get(k) for k in MET_KEYS if src.get(k) is not None}


def all_rows():
    for name in ('block1', 'newItem', 'newCat', 'runItem', 'runCat', 'seasonCat'):
        for d in SUM[name]:
            yield d


SNAP = os.path.join(HERE, 'sum_prev.json')
snap = json.load(open(SNAP, encoding='utf-8')) if os.path.exists(SNAP) else None
# v3: 스타일수/SKU/기획량 정의가 MDP 기준으로 바뀜 — 구버전 값은 비교하지 않는다
if snap and snap.get('ver') != 3:
    for _rows in (snap.get('rows', {}), snap.get('carry', {})):
        for _v in _rows.values():
            for _k in ('styles', 'sku', 'plan', 'inRate'):
                _v.pop(_k, None)
# ---- 스타일수/SKU/기획량은 MDP(수량 기획서)에서, 재고량은 카드 데이터에서 집계해 SUM 행에 주입 ----
from collections import defaultdict as _dd

# 재고: 신상품 = 입고-누계판매(카드), 러닝 = 시점재고+신규입고(카드 stock)
ITEM2BOK = {'OUTER': 'OUTER', 'KNIT': 'TOP', 'INNER (W/O KNIT)': 'TOP',
            'BOTTOM': 'BOTTOM', 'ACC': 'ACC', '기타': 'ACC'}
stkB, stkC = _dd(float), _dd(float)
for st in data:
    bok = ITEM2BOK.get(st['item'], 'ACC')
    if st['line'] == '메인':
        v = (st['ttl']['qtyIn'] or 0) - (st['ttl']['total'] or 0)
        stkB[(st['season'], bok)] += v
        stkC[(st['season'], st['cat2'])] += v
    else:
        v = st.get('stock', 0)
        stkB[('러닝', bok)] += v
        stkC[('러닝', st['cat2'])] += v

# MDP: data/26FW_MDP.xlsb 의 MDP 시트를 스타일 블록(AI열 STY=1 이 시작) 단위로 읽어
# ED열(컨펌여부)이 '확정' 인 스타일만 집계한다. SKU = 블록 안 컬러명(AM) 행 수,
# 기획량 = 컬러별수량(AN) 합. 구분 매핑: INNER -> TOP, SHOES -> ACC.
MDP_PATH = os.path.join(DATA_DIR, '26FW_MDP.xlsb')
mdpSeason, mdpBok = {}, {}
mdp_loaded = False
if os.path.exists(MDP_PATH):
    from pyxlsb import open_workbook as _oxb
    _iE, _iF, _iH, _iAI, _iAJ, _iAK, _iAM, _iAN, _iED = (
        C('E'), C('F'), C('H'), C('AI'), C('AJ'), C('AK'), C('AM'), C('AN'), C('ED'))
    BOKMAP = {'OUTER': 'OUTER', 'INNER': 'TOP', 'BOTTOM': 'BOTTOM', 'ACC': 'ACC', 'SHOES': 'ACC'}
    blocks, cur = [], None
    with _oxb(MDP_PATH) as _wb, _wb.get_sheet('MDP') as _ws:
        for _row in _ws.rows():
            _d = {c.c: c.v for c in _row}
            if _d.get(_iAI) == 1:                      # 스타일 시작
                if cur:
                    blocks.append(cur)
                _season = txt(_d.get(_iE))
                # SKU/기획량은 스타일 행의 공식 열(AJ/AK)을 쓴다 — 컬러 행(AM/AN)은
                # 미입력분이 있어 합계가 모자랄 수 있다 (예: 컬러명 없는 SKU, 수량 미배분)
                cur = {'season': '러닝' if '(러닝)' in _season else _season,
                       'bok': BOKMAP.get(txt(_d.get(_iF)), 'ACC'),
                       'code': txt(_d.get(_iH)),                     # 아이템코드 (집계표 카테고리 코드와 대응)
                       'ok': False,
                       'sku': _d.get(_iAJ) if isinstance(_d.get(_iAJ), (int, float)) else 0,
                       'plan': _d.get(_iAK) if isinstance(_d.get(_iAK), (int, float)) else 0.0}
            if cur is None:
                continue
            if txt(_d.get(_iED)) == '확정':
                cur['ok'] = True
    if cur:
        blocks.append(cur)
    mdpCat, _code_fbok = {}, {}
    for b in blocks:
        if not b['ok'] or b['season'] not in ('가을', '겨울', '러닝'):
            continue
        _code_fbok.setdefault(b['code'], b['bok'])
        for tgt in (mdpSeason.setdefault(b['season'], {'styles': 0, 'sku': 0, 'plan': 0}),
                    mdpCat.setdefault((b['season'], b['code']), {'styles': 0, 'sku': 0, 'plan': 0})):
            tgt['styles'] += 1
            tgt['sku'] += b['sku']
            tgt['plan'] += b['plan']
    # 복종 합계는 화면(종합 시트) 그룹 기준으로 아이템코드를 다시 묶는다.
    # MDP F열과 그룹이 다른 코드(예: 니트베스트 KV — MDP는 INNER, 화면은 OUTER)가 있어
    # F열로 합치면 부모 행과 하위 카테고리 합이 어긋난다.
    code2grp = {}
    for _name in ('seasonCat', 'newCat', 'runCat'):
        for _d in SUM[_name]:
            g = _d['item']
            code2grp.setdefault(_d['code'], 'ACC' if g.startswith('ACC') else g)
    for (season, code), v in mdpCat.items():
        g = code2grp.get(code) or _code_fbok.get(code, 'ACC')
        t = mdpBok.setdefault((season, g), {'styles': 0, 'sku': 0, 'plan': 0})
        for k in t:
            t[k] += v[k]
    mdpSeason['TTL'] = {k: sum(v[k] for v in mdpSeason.values()) for k in ('styles', 'sku', 'plan')}
    mdp_loaded = True
    print('MDP 확정 기준:', {k: (v['styles'], v['sku'], round(v['plan'])) for k, v in mdpSeason.items()},
          '| 미확정 제외', sum(1 for b in blocks if not b['ok']), '스타일')
else:
    print('경고: %s 없음 — 스타일수/SKU/기획량을 MDP 로 대체하지 못함' % MDP_PATH)


def _sumv(ds):
    out = {'styles': 0, 'sku': 0, 'plan': 0}
    for d in ds:
        for k in out:
            out[k] += d[k]
    return out


def _put(row, agg, stock):
    if mdp_loaded:
        for k in ('styles', 'sku', 'plan'):
            row[k] = agg[k] if agg else None
        row.pop('inRate', None)          # 기획량이 MDP 기준으로 바뀌었으니 입고율은 다시 계산
    row['stock'] = stock


_Z = {'styles': 0, 'sku': 0, 'plan': 0}
_seasons = ('가을', '겨울')
_newTtl = _sumv(mdpSeason.get(k, _Z) for k in _seasons) if mdp_loaded else None
for d in SUM['block1']:
    if d['mode'] == 'run':
        continue
    tgt = d['cur']
    if d.get('kind') == '신상품 TTL':
        _put(tgt, _newTtl, sum(v for k, v in stkB.items() if k[0] in _seasons))
    elif d['item'] == 'TTL':
        _put(tgt, mdpSeason.get(d['season']), sum(v for k, v in stkB.items() if k[0] == d['season']))
    else:
        _put(tgt, mdpBok.get((d['season'], d['item'])), stkB[(d['season'], d['item'])])
for d in SUM['newItem']:
    if d['item'] == 'TOTAL':
        _put(d, _newTtl, sum(v for k, v in stkB.items() if k[0] in _seasons))
    else:
        _put(d, _sumv(mdpBok.get((se, d['item']), _Z) for se in _seasons),
             sum(v for k, v in stkB.items() if k[0] in _seasons and k[1] == d['item']))
for d in SUM['runItem']:
    if d['item'] == 'TOTAL':
        _put(d, mdpSeason.get('러닝'), sum(v for k, v in stkB.items() if k[0] == '러닝'))
    else:
        _put(d, mdpBok.get(('러닝', d['item'])), stkB[('러닝', d['item'])])
# 하위 카테고리: MDP 확정 스타일을 아이템코드별로 집계해 채운다
if mdp_loaded:
    _Zc = {'styles': 0, 'sku': 0, 'plan': 0}
    for d in SUM['seasonCat']:
        _put(d, mdpCat.get((d['season'], d['code']), _Zc), d.get('stock', 0))
    for d in SUM['newCat']:
        _put(d, _sumv(mdpCat.get((se, d['code']), _Zc) for se in _seasons), d.get('stock', 0))
    for d in SUM['runCat']:
        _put(d, mdpCat.get(('러닝', d['code']), _Zc), d.get('stock', 0))

print('집계표:', {k: len(v) for k, v in SUM.items()})


# 업데이트 날짜: 파일명 앞 6자리(YYMMDD)
base = os.path.basename(SRC)
updated = ''
if base[:6].isdigit():
    updated = '20%s-%s-%s' % (base[0:2], base[2:4], base[4:6])
print('updated:', updated)

if snap and snap.get('date') and snap['date'] != updated:
    hit = 0
    for d in all_rows():
        was = snap['rows'].get(d['_k'])
        if was:
            d['was'] = was
            hit += 1
    print('증감 비교: %s -> %s (%d행)' % (snap['date'], updated, hit))
elif snap:
    # 같은 날짜 파일을 다시 빌드한 경우 — 직전 스냅샷을 그대로 두고 증감도 유지
    for d in all_rows():
        was = snap.get('carry', {}).get(d['_k'])
        if was:
            d['was'] = was
    if snap.get('carry'):
        print('증감 비교: %s 유지' % snap.get('carryDate'))
else:
    print('증감 비교: 직전 스냅샷 없음 (이번 빌드부터 기록)')

# 스냅샷은 매 빌드 다시 쓴다 (지표 정의가 바뀌어도 다음 비교가 새 정의로 이뤄지도록).
# 같은 날짜를 다시 빌드하면 carry(직전 날짜 값)는 그대로 두어 증감 기준이 유지된다.
if not snap or snap.get('date') != updated:
    carry, carry_date = (snap or {}).get('rows', {}), (snap or {}).get('date')
else:
    carry, carry_date = snap.get('carry', {}), snap.get('carryDate')
json.dump({'ver': 3, 'date': updated,
           'rows': {d['_k']: row_vals(d) for d in all_rows()},
           'carry': carry, 'carryDate': carry_date},
          open(SNAP, 'w', encoding='utf-8'), ensure_ascii=False)
print('스냅샷 저장:', os.path.basename(SNAP), updated)

def render(tpl_name, out_path, **repl):
    tpl = open(os.path.join(HERE, tpl_name), encoding='utf-8').read()
    for k, v in repl.items():
        tpl = tpl.replace(k, v)
    tpl = tpl.replace('__UPDATED__', updated)
    open(out_path, 'w', encoding='utf-8').write(tpl)
    print('written:', out_path, round(len(tpl.encode()) / 1e6, 2), 'MB')


render('salecards_template.html', OUT_CARD,
       **{'/*__DATA__*/[]': json.dumps(data, ensure_ascii=False)})
render('summary_template.html', OUT_SUM,
       **{'/*__SUM__*/{}': json.dumps(SUM, ensure_ascii=False)})
